from openpyxl import Workbook


def edit(word):
    if '.cpp' in word:
        word = word.split(':')
    elif '.h' in word:
        word = word.split(':')
    elif keyword in word:
        word = word.replace(":", "")
    elif 'sc' in word:
        word = word.replace(":", "")
    elif 'rb' in word:
        word = word.replace(",", " ")
        word = word.replace(":", " ")
        word = word.split()
    elif "," in word:
        word = word.replace(",", " ")
    elif '[' in word:
        word = word.split()

    return word


def write_all(write_ws):
    cppnames = set()
    i = 1
    for line in keylines:
        words = line.split(maxsplit=7)
        j = 1
        cpp = words[5].split(':')
        cppnames.add(cpp[0])

        for word in words:
            word = edit(word)
            if type(word) is list:
                for w in word:
                    write_ws.cell(row=i, column=j).value = w
                    j += 1
            else:
                write_ws.cell(row=i, column=j).value = word
                j += 1
        i += 1
    return cppnames


def write_cpp(write_wb, cppnames):
    global keylines, keyword
    for cpp in cppnames:
        # make sheet
        write_ws = write_wb.create_sheet(cpp)
        i = 1
        print(keylines)
        for line in keylines:
            words = line.split(maxsplit=7)
            j = 1
            if cpp in line:
                for word in words:
                    word = edit(word)
                    if type(word) is list:
                        for w in word:
                            write_ws.cell(row=i, column=j).value = w
                            j += 1
                    else:
                        write_ws.cell(row=i, column=j).value = word
                        j += 1
                i += 1


def write_rb(wb):
    length = 0
    rb = start_rb
    time = slot-1
    wb.remove(wb['Sheet'])
    while length != len(lines):
        if length % (14*12) == 0:
            time += 1
            ws = wb.create_sheet(str(time))
            row_num = 1

        ws.cell(row=row_num, column=1).value = "rb" + str(rb)
        for i in range(1, 7):
            words = lines[length].split()
            j = 0
            for word in words:
                ws.cell(row=row_num, column=i * 2 + j).value = word
                j = j + 1
            length += 1
        row_num += 1
        rb += 1
        if rb == (start_rb + num_rb):
            rb = start_rb


def excel():
    global lines
    # write file
    wb = Workbook()

    if sel == "1":
        write_rb(wb)
        wb.save('0201_iq.xlsx')
    elif sel == "2":
        write_ws = wb.active
        write_ws.title = 'total'
        cppnames = write_all(write_ws)
        write_cpp(wb, cppnames)
        wb.save('0201_log.xlsx')


def getfile():
    global lines, keylines
    f = open(filepath, 'r')
    lines = f.readlines()
    for line in lines:
        if keyword in line:
            keylines.append(line)
    f.close()


def iqprint():
    global lines, start_rb, num_rb, slot
    start_rb = int(input("Write start rb"))
    num_rb = int(input("Write number of rb"))
    slot = int(input("Write start of slot"))
    f = open(filepath, 'r')
    lines = f.readlines()
    f.close()


def getpath():
    global filepath, keyword, sel
    print("Place your data in the same directory with this file")
    # excelpath = input('Write full path to new excel:\n')
    # filepath = input('Write file name of your data:\n')
    # keyword = input('Write your keyword. It will print the line that contains the keyword')
    # fkey = keyword.replace("/")
    # print(filepath)
    # filepath.replace('\\', '/')

    sel = input("1:iqsample 2:log 3: rawdata")
    if sel == "1":
        filepath = 'iqsample'
        iqprint()
    elif sel == "2":
        keyword = 'I/PUSCH'
        filepath = 'mlog_latest'
        getfile()
    elif sel == "3":
        return


lines = []
keylines = []
filepath = ""
keyword = ""
sel = ""
start_rb = num_rb = slot = 0

getpath()
excel()
